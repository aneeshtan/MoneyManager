#!/usr/bin/env python3
import argparse
import json
import re
import sqlite3
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import to_excel


LOCATION_SUFFIXES = (
    "DUBAI ARE",
    "DUBAI AE",
    "DUBAI",
    "ABU DHABI ARE",
    "ABU DHABI AE",
    "ABU DHABI",
    "RAS AL KHAIMA",
    "RAS AL-KHAIMA",
    "AJMAN",
    "SHARJAH",
    "UAE",
    "AE",
)


def clean_name(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_merchant(value):
    text = clean_name(value).upper()
    text = re.sub(r"CRCARDXXX\d+USED", "", text)
    text = re.sub(r"CARDXXX\d+USED", "", text)
    text = re.sub(r"\(\+\d+(?:\.\d+)?%[^)]*\)", "", text)
    text = re.sub(r"AVL\.?\s*CR\.?\s*LIMIT(?:\s*IS)?\s*AED[\d,.]+", "", text)
    text = re.sub(r"AVLCRLIMIT(?:IS)?AED[\d,.]+", "", text)
    text = re.sub(r"[^A-Z0-9&%+./ -]", " ", text)
    text = re.sub(r"\s+", " ", text).strip(" -,.")
    changed = True
    while changed:
        changed = False
        for suffix in LOCATION_SUFFIXES:
            if text.endswith(" " + suffix):
                text = text[: -len(suffix)].strip(" -,.")
                changed = True
    return text


def read_accounts(db_path):
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    rows = con.execute(
        """
        SELECT ZUID, ZNICNAME, ZCURRENCYUID, ZORDER
        FROM ZASSET
        WHERE ZISDEL = 0
        ORDER BY ZORDER, ZNICNAME
        """
    ).fetchall()
    con.close()
    return [
        {
            "id": clean_name(row["ZUID"]),
            "name": clean_name(row["ZNICNAME"]),
            "currency": currency_from_uid(clean_name(row["ZCURRENCYUID"])),
            "sortOrder": int(row["ZORDER"] or index),
        }
        for index, row in enumerate(rows)
        if clean_name(row["ZNICNAME"])
    ]


def currency_from_uid(uid):
    if uid.endswith("_USD"):
        return "USD"
    return "AED"


def read_categories(db_path):
    con = sqlite3.connect(db_path)
    con.row_factory = sqlite3.Row
    rows = con.execute(
        """
        SELECT ZUID, ZNAME, ZDOTYPE, ZPUID, ZORDER
        FROM ZCATEGORY
        WHERE ZISDEL = 0
        ORDER BY ZDOTYPE, ZORDER, ZNAME
        """
    ).fetchall()
    con.close()
    categories = []
    for row in rows:
        name = clean_name(row["ZNAME"])
        if not name:
            continue
        parent = clean_name(row["ZPUID"])
        categories.append(
            {
                "id": clean_name(row["ZUID"]),
                "name": name,
                "kind": "income" if int(row["ZDOTYPE"] or 0) == 0 else "expense",
                "parentId": None if parent in ("", "0") else parent,
                "sortOrder": int(row["ZORDER"] or 0),
            }
        )
    return categories


def read_merchant_rules(xlsx_path):
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active
    votes = defaultdict(Counter)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[4] or not row[2]:
            continue
        merchant = normalize_merchant(row[4])
        if len(merchant) < 3:
            continue
        category = clean_name(row[2])
        subcategory = clean_name(row[3])
        kind = "income" if clean_name(row[6]).lower().startswith("income") else "expense"
        votes[merchant][(category, subcategory, kind)] += 1

    rules = []
    for merchant, counts in votes.items():
        (category, subcategory, kind), count = counts.most_common(1)[0]
        total = sum(counts.values())
        if count < 2 and total < 3:
            continue
        confidence = count / total
        rules.append(
            {
                "pattern": merchant,
                "matchType": "contains",
                "category": category,
                "subcategory": subcategory or None,
                "kind": kind,
                "confidence": round(confidence, 3),
                "sampleCount": int(total),
            }
        )

    rules.extend(fallback_rules())
    rules.sort(key=lambda item: (-item["sampleCount"], item["pattern"]))
    return rules


def read_initial_transactions(xlsx_path):
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active
    transactions = []
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        if not row or not row[0]:
            continue
        period = row[0]
        if isinstance(period, datetime):
            date_iso = period.isoformat()
            period_serial = decimal_string(to_excel(period))
        else:
            date_iso = clean_name(period)
            period_serial = decimal_string(period)

        account = clean_name(row[1])
        category = clean_name(row[2])
        subcategory = clean_name(row[3])
        note = clean_name(row[4])
        income_expense = clean_name(row[6])
        description = clean_name(row[7])
        currency = clean_name(row[9]) or "AED"
        amount = decimal_string(row[8] if len(row) > 8 else row[5])
        aed_amount = decimal_string(row[5])
        trailing_accounts = decimal_string(row[10]) if len(row) > 10 else ""
        merchant = note or description or category or "Imported transaction"

        transactions.append(
            {
                "sourceRow": index,
                "date": date_iso,
                "periodSerial": period_serial,
                "account": account,
                "category": category,
                "subcategory": subcategory or None,
                "note": note,
                "aed": aed_amount,
                "incomeExpense": income_expense,
                "description": description,
                "amount": amount,
                "currency": currency,
                "accountsTrailing": trailing_accounts,
                "kind": kind_from_export(income_expense),
                "merchant": merchant,
                "normalizedMerchant": normalize_merchant(merchant),
            }
        )
    return transactions


def decimal_string(value):
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.10f}".rstrip("0").rstrip(".")
    return clean_name(value)


def kind_from_export(value):
    text = clean_name(value).lower()
    if text.startswith("income"):
        return "income"
    if text.startswith("transfer"):
        return "transfer"
    return "expense"


def fallback_rules():
    return [
        rule("CAREEM HALA", "Transportation", "Taxi"),
        rule("CAREEM RIDE", "Transportation", "Taxi"),
        rule("NATIONAL TAXI", "Transportation", "Taxi"),
        rule("DUBAI TAXI", "Transportation", "Taxi"),
        rule("AMAZON GROCERY", "Food", None),
        rule("AMAZON NOW", "Food", None),
        rule("UNION COOP", "Food", None),
        rule("AL MAYA", "Food", None),
        rule("CARREFOUR", "Food", None),
        rule("LAUNDRY", "Apparel", "Laundry"),
        rule("AGODA", "Travel", None),
        rule("BOOKING", "Travel", None),
        rule("ADNOC", "Car", "Petrol"),
        rule("EMARAT", "Car", "Petrol"),
        rule("SALIK", "Car", "Salik"),
        rule("DU APPLE PAY", "Internet", None),
        rule("IKEA", "Household", None),
        rule("JUSTLIFE", "Household", None),
    ]


def rule(pattern, category, subcategory, kind="expense"):
    return {
        "pattern": pattern,
        "matchType": "contains",
        "category": category,
        "subcategory": subcategory,
        "kind": kind,
        "confidence": 0.75,
        "sampleCount": 0,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--backup", required=True, type=Path)
    parser.add_argument("--xlsx", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--display-name", default="User")
    parser.add_argument("--default-import-account", default="Credit Card")
    parser.add_argument("--include-transactions", action="store_true")
    args = parser.parse_args()

    seed = {
        "version": 1,
        "user": {
            "id": "primary-user",
            "displayName": args.display_name,
            "baseCurrency": "AED",
        },
        "defaultCurrency": "AED",
        "defaultImportAccount": args.default_import_account,
        "accounts": read_accounts(args.backup),
        "categories": read_categories(args.backup),
        "merchantRules": read_merchant_rules(args.xlsx),
        "initialTransactions": read_initial_transactions(args.xlsx) if args.include_transactions else [],
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(seed, ensure_ascii=False, indent=2) + "\n")
    print(
        f"Wrote {len(seed['accounts'])} accounts, "
        f"{len(seed['categories'])} categories, "
        f"{len(seed['merchantRules'])} merchant rules, "
        f"{len(seed['initialTransactions'])} transactions to {args.output}"
    )


if __name__ == "__main__":
    main()
